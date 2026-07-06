"""Tabular export helpers for reports (Excel via openpyxl, PDF via reportlab)."""
from __future__ import annotations

import io
from typing import Any, Sequence


def to_excel(title: str, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> bytes:
    """Render a simple table to an .xlsx workbook and return its bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"  # Excel sheet name limit is 31 chars.

    # Title row.
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # Header row.
    ws.append(list(headers))
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows.
    for row in rows:
        ws.append([_cell(v) for v in row])

    # Auto-ish column widths.
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[_col_letter(col_idx)].width = min(max_len + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def to_pdf(title: str, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> bytes:
    """Render a simple table to a PDF and return its bytes."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements: list[Any] = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

    data: list[list[str]] = [list(headers)]
    for row in rows:
        data.append([str(_cell(v)) for v in row])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E1F2")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()


def _cell(value: Any) -> Any:
    """Normalise a value for output (Decimals -> float, None -> '')."""
    from decimal import Decimal

    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    return value


def _col_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(index)
